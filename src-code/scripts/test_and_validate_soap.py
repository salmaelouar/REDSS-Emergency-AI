#!/usr/bin/env python3
"""
Script to:
1. Clear all existing emergency call data
2. Import 15 test cases from evaluated_text_data.py
3. Generate a PDF report with S and O summaries for nurse validation
"""

import sys
import os
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import sqlite3
import json
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import pandas as pd

# Database paths
DB_PATH = project_root / "data" / "emergency_calls.db"
PATIENT_DB_PATH = project_root / "data" / "registered_patients.db"
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
PDF_OUTPUT = project_root / "reports" / f"SOAP_Validation_Report_{timestamp}.pdf"
EXCEL_OUTPUT = project_root / "reports" / f"SOAP_Validation_Report_{timestamp}.xlsx"

def clear_all_data():
    """Clear all data from both databases"""
    print("🗑️  Step 1: Clearing all existing data...")
    
    # Clear emergency calls database
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM emergency_calls")
        deleted_calls = cursor.rowcount
        conn.commit()
        conn.close()
        print(f"   ✓ Deleted {deleted_calls} emergency calls")
    except Exception as e:
        print(f"   ⚠️  Error clearing emergency calls: {e}")
    
    # Clear patient database
    try:
        conn = sqlite3.connect(PATIENT_DB_PATH)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM registered_patients")
        deleted_patients = cursor.rowcount
        conn.commit()
        conn.close()
        print(f"   ✓ Deleted {deleted_patients} registered patients")
    except Exception as e:
        print(f"   ⚠️  Error clearing patients: {e}")
    
    print("   ✅ Database cleared successfully!\n")

def import_test_cases():
    """Import 15 test cases and process them through the pipeline"""
    print("📥 Step 2: Importing 15 test cases...")
    
    # Import the test data
    sys.path.insert(0, str(project_root / "data" / "text"))
    
    # Read the evaluated_text_data.py file as a Python list
    with open(project_root / "data" / "text" / "evaluated_text_data.py", 'r') as f:
        content = f.read()
        # Execute the content to get the list
        test_data = eval(content)
    
    # Initialize the pipeline
    from app.services.pipeline import pipeline
    from app.services.database import get_db
    from app.models.call import EmergencyCall
    
    processed_calls = []
    
    with get_db() as db:
        for i, call_data in enumerate(test_data, 1):
            print(f"   Processing {i}/15: {call_data['call_id']}...")
            
            try:
                # Process through pipeline
                result = pipeline.process_text(
                    transcript=call_data['text'],
                    patient_name=call_data.get('expected_agent', ''),
                    language='en'
                )
                
                # Get the processed call from database
                call = db.query(EmergencyCall).filter(
                    EmergencyCall.call_id == result['call_id']
                ).first()
                
                if call:
                    processed_calls.append({
                        'call_id': call_data['call_id'],
                        'db_call_id': call.call_id,
                        'patient_name': call.patient_name,
                        'urgency_level': call.urgency_level,
                        'urgency_reasoning': call.urgency_reasoning,
                        'expected_urgency': call_data.get('expected_urgency', '').upper(),
                        'subjective': call.soap_subjective,
                        'objective': call.soap_objective,
                        'assessment': call.soap_assessment,
                        'plan': call.soap_plan,
                        'expected_subjective': call_data['expected_soap']['subjective'],
                        'expected_objective': call_data['expected_soap']['objective'],
                    })
                    print(f"      ✓ Processed successfully")
                
            except Exception as e:
                print(f"      ✗ Error: {e}")
                continue
    
    print(f"   ✅ Imported and processed {len(processed_calls)}/15 calls\n")
    return processed_calls

def generate_pdf_report(processed_calls):
    """Generate PDF report with S and O summaries for nurse validation"""
    print("📄 Step 3: Generating PDF validation report...")
    
    # Ensure reports directory exists
    PDF_OUTPUT.parent.mkdir(exist_ok=True)
    
    # Create PDF
    doc = SimpleDocTemplate(
        str(PDF_OUTPUT),
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.5*inch,
    )
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a237e'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#283593'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    subheading_style = ParagraphStyle(
        'CustomSubHeading',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor('#3949ab'),
        spaceAfter=6,
        spaceBefore=6,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        spaceAfter=6
    )
    
    small_style = ParagraphStyle(
        'SmallText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.grey
    )
    
    # Title Page
    elements.append(Spacer(1, 1*inch))
    elements.append(Paragraph("SOAP NOTES & TRIAGE VALIDATION REPORT", title_style))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(
        f"AI-Extracted Clinical Notes & Urgency Classification",
        ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=14, alignment=TA_CENTER, textColor=colors.grey)
    ))
    elements.append(Spacer(1, 0.3*inch))
    
    # Info box
    info_data = [
        ['Report Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total Cases:', str(len(processed_calls))],
        ['Purpose:', 'Clinical validation of AI extractions'],
        ['Focus:', 'SOAP notes (S & O) + Urgency triage'],
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8eaf6')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#283593')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#c5cae9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Instructions for nurse
    elements.append(Paragraph("VALIDATION INSTRUCTIONS", heading_style))
    instructions = """
    Dear Nurse Reviewer,<br/><br/>
    
    Please review each case and validate the following:<br/>
    • <b>S + O Summary</b>: Are the patient's complaints, symptoms, and observable signs accurately captured?<br/>
    • <b>Triage/Urgency</b>: Is the urgency level classification appropriate for the case?<br/><br/>
    
    For each section, please mark:<br/>
    ✓ <b>ACCURATE</b> - Information is correctly extracted/classified<br/>
    ⚠ <b>INCOMPLETE</b> - Missing important information<br/>
    ✗ <b>INCORRECT</b> - Contains errors or misinterpretations<br/><br/>
    
    Your feedback will help improve both AI extraction accuracy and triage logic.
    """
    elements.append(Paragraph(instructions, normal_style))
    
    elements.append(PageBreak())
    
    # Process each call
    for idx, call in enumerate(processed_calls, 1):
        # Case header with colored background
        case_header = Table(
            [[f"Case {idx} of {len(processed_calls)}: {call['call_id']}", f"Urgency: {call['urgency_level']}"]],
            colWidths=[5*inch, 2*inch]
        )
        case_header.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#3f51b5')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(case_header)
        elements.append(Spacer(1, 0.15*inch))
        
        # Patient name
        if call.get('patient_name'):
            elements.append(Paragraph(f"<b>Patient:</b> {call['patient_name']}", normal_style))
            elements.append(Spacer(1, 0.1*inch))
        
        # ========== TRIAGE/URGENCY SECTION ==========
        elements.append(Paragraph("AI TRIAGE CLASSIFICATION", subheading_style))
        
        ai_urg = call.get('urgency_level', 'UNKNOWN')
        
        # Color coding for urgency levels
        urgency_colors = {
            'CRITICAL': colors.HexColor('#d32f2f'),
            'HIGH': colors.HexColor('#f57c00'),
            'MEDIUM': colors.HexColor('#fbc02d'),
            'LOW': colors.HexColor('#388e3c'),
        }
        
        # Simple triage display - AI Predicted only
        triage_box = Table(
            [[Paragraph(f"<b>AI Predicted Urgency: {ai_urg}</b>", normal_style)]],
            colWidths=[7*inch]
        )
        triage_box.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), urgency_colors.get(ai_urg, colors.lightgrey)),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.white),
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
            ('TOPPADDING', (0, 0), (0, 0), 12),
            ('BOTTOMPADDING', (0, 0), (0, 0), 12),
        ]))
        elements.append(triage_box)
        elements.append(Spacer(1, 0.1*inch))
        
        # AI Reasoning - Why this classification?
        if call.get('urgency_reasoning'):
            reasoning_text = f"<b>Triage Logic:</b> {call['urgency_reasoning']}"
            reasoning_box = Table(
                [[Paragraph(reasoning_text, normal_style)]],
                colWidths=[7*inch]
            )
            reasoning_box.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#fffde7')),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#fdd835')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(reasoning_box)
            elements.append(Spacer(1, 0.1*inch))
        
        # Nurse Validation
        validation_triage = Table(
            [[Paragraph("<b>Nurse Assessment:</b>", normal_style)], ['☐ ACCURATE', '☐ SHOULD BE HIGHER', '☐ SHOULD BE LOWER', 'Comments: __________________']],
            colWidths=[1.2*inch, 1.6*inch, 1.6*inch, 2.6*inch]
        )
        validation_triage.setStyle(TableStyle([
            ('SPAN', (0, 0), (3, 0)),
            ('BACKGROUND', (0, 0), (3, 0), colors.HexColor('#e8eaf6')),
            ('FONTNAME', (0, 0), (3, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (3, 0), 10),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, 1), 9),
            ('ALIGN', (0, 1), (2, 1), 'CENTER'),
            ('ALIGN', (3, 1), (3, 1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(validation_triage)
        elements.append(Spacer(1, 0.25*inch))
        

        # S + O COMBINED Section - Plain merged text
        elements.append(Paragraph("CLINICAL SUMMARY (S + O)", subheading_style))
        
        # Merge Subjective and Objective as plain continuous text (no labels)
        combined_text = ""
        if call.get('subjective') and call.get('objective'):
            # Merge seamlessly - subjective flows into objective
            combined_text = f"{call['subjective']} {call['objective']}"
        elif call.get('subjective'):
            combined_text = call['subjective']
        elif call.get('objective'):
            combined_text = call['objective']
        else:
            combined_text = "[No clinical summary extracted]"
        
        # Display as simple text box
        so_box = Table(
            [[Paragraph(combined_text, normal_style)]],
            colWidths=[7*inch]
        )
        so_box.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, 0), colors.white),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#9fa8da')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(so_box)
        elements.append(Spacer(1, 0.1*inch))
        
        # Single validation checkbox for S+O
        validation_so = Table(
            [['☐ ACCURATE', '☐ INCOMPLETE', '☐ INCORRECT', 'Comments: _________________________']],
            colWidths=[1.2*inch, 1.3*inch, 1.2*inch, 3.3*inch]
        )
        validation_so.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (2, 0), 'CENTER'),
            ('ALIGN', (3, 0), (3, 0), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(validation_so)
        elements.append(Spacer(1, 0.2*inch))
        
        # Optional: Show Assessment and Plan for context
        elements.append(Paragraph("<i>For Context:</i>", small_style))
        context = f"""
        <b>Assessment:</b> {call.get('assessment', 'N/A')}<br/>
        <b>Plan:</b> {call.get('plan', 'N/A')}
        """
        elements.append(Paragraph(context, small_style))
        
        # ===== ONE CASE PER PAGE - Always add page break after each case =====
        elements.append(PageBreak())
        
 # Summary page with triage statistics
    elements.append(Paragraph("TRIAGE ACCURACY SUMMARY", heading_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Calculate matches and mismatches
    total_cases = len(processed_calls)
    matches = sum(1 for call in processed_calls if call.get('expected_urgency') == call.get('urgency_level'))
    mismatches = total_cases - matches
    accuracy_pct = (matches / total_cases * 100) if total_cases > 0 else 0
    
    # Count by urgency level
    urgency_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
    for call in processed_calls:
        level = call.get('urgency_level', 'UNKNOWN')
        if level in urgency_counts:
            urgency_counts[level] += 1
    
    # Overview stats box
    stats_box_data = [
        ['Total Cases Analyzed:', str(total_cases)],
        ['Expected vs AI Matches:', f"{matches} / {total_cases}"],
        ['Mismatches:', str(mismatches)],
        ['AI Accuracy Rate:', f"{accuracy_pct:.1f}%"],
    ]
    
    stats_box = Table(stats_box_data, colWidths=[3*inch, 2*inch])
    stats_box.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f5f5f5')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(stats_box)
    elements.append(Spacer(1, 0.3*inch))
    
    # AI Triage Distribution
    elements.append(Paragraph("AI Triage Distribution", subheading_style))
    distribution_data = [
        ['Urgency Level', 'Count', 'Percentage'],
        ['CRITICAL', str(urgency_counts['CRITICAL']), f"{urgency_counts['CRITICAL']/total_cases*100:.1f}%" if total_cases > 0 else "0%"],
        ['HIGH', str(urgency_counts['HIGH']), f"{urgency_counts['HIGH']/total_cases*100:.1f}%" if total_cases > 0 else "0%"],
        ['MEDIUM', str(urgency_counts['MEDIUM']), f"{urgency_counts['MEDIUM']/total_cases*100:.1f}%" if total_cases > 0 else "0%"],
        ['LOW', str(urgency_counts['LOW']), f"{urgency_counts['LOW']/total_cases*100:.1f}%" if total_cases > 0 else "0%"],
    ]
    
    dist_table = Table(distribution_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    dist_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3f51b5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ffcdd2')),  # CRITICAL
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#ffecb3')),  # HIGH
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fff9c4')),  # MEDIUM
        ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#c8e6c9')),  # LOW
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(dist_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Mismatch list (if any)
    if mismatches > 0:
        elements.append(Paragraph(f"Cases with Expected vs AI Mismatch ({mismatches} cases):", subheading_style))
        mismatch_data = [['Case ID', 'Expected', 'AI Predicted']]
        for call in processed_calls:
            if call.get('expected_urgency') != call.get('urgency_level'):
                mismatch_data.append([
                    call['call_id'],
                    call.get('expected_urgency', 'N/A'),
                    call.get('urgency_level', 'N/A')
                ])
        
        mismatch_table = Table(mismatch_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        mismatch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ff5722')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(mismatch_table)
        elements.append(Spacer(1, 0.3*inch))
    

    # Reviewer signature section - ON SEPARATE PAGE
    elements.append(PageBreak())
    elements.append(Paragraph("Clinical Reviewer Information", heading_style))
    summary_data = [
        ['Reviewer Name:', '_________________________________'],
        ['Professional Title:', '_________________________________'],
        ['Review Date:', '_________________________________'],
        ['Signature:', '_________________________________'],
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 3.5*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('LINEBELOW', (1, 0), (1, -1), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Additional comments section
    elements.append(Paragraph("Overall Assessment & Recommendations:", heading_style))
    elements.append(Spacer(1, 0.1*inch))
    
    for i in range(8):
        elements.append(Paragraph("_" * 120, normal_style))
        

    # --- APPENDIX: SYSTEM LOGIC EXPLANATION ---
    elements.append(PageBreak())
    elements.append(Paragraph("APPENDIX: SYSTEM LOGIC EXPLANATION", heading_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # 1. Scientific Basis
    elements.append(Paragraph("1. SCIENTIFIC BASIS", subheading_style))
    elements.append(Paragraph("<b>Reference:</b> Gilboy N, Tanabe P, Travers D, Rosenau AM. Emergency Severity Index (ESI): A Triage Tool for Emergency Department Care, Version 4. Implementation Handbook 2012 Edition. AHRQ Publication No. 12-0014. Rockville, MD.", normal_style))
    elements.append(Spacer(1, 0.15*inch))
    
    # 2. ESI Standard Levels Table
    elements.append(Paragraph("2. ESI STANDARD LEVELS (Medical Standard)", subheading_style))
    esi_data = [
        ['Level', 'Name', 'Description'],
        ['1', 'Resuscitation', 'Immediate life-saving intervention needed (e.g. Cardiac Arrest)'],
        ['2', 'Emergent', 'High risk, confused, severe pain/distress (e.g. Chest Pain, Stroke)'],
        ['3', 'Urgent', 'Stable, multiple resources needed (e.g. Abdominal Pain)'],
        ['4', 'Less Urgent', 'Stable, one resource needed (e.g. Simple Laceration)'],
        ['5', 'Non-Urgent', 'No resources needed (e.g. Prescription Refill)']
    ]
    esi_table = Table(esi_data, colWidths=[0.6*inch, 1.2*inch, 4.5*inch])
    esi_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e0e0')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(esi_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 3. Mapping to Dashboard
    elements.append(Paragraph("3. MAPPING TO DASHBOARD LEVELS", subheading_style))
    map_data = [
        ['ESI Level', 'Dashboard Urgency Level', 'Color'],
        ['Level 1', 'CRITICAL', 'RED'],
        ['Level 2', 'HIGH PRIORITY', 'ORANGE'],
        ['Level 3', 'MEDIUM PRIORITY', 'YELLOW'],
        ['Level 4 & 5', 'LOW PRIORITY', 'GREEN']
    ]
    map_table = Table(map_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch])
    map_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 1, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e0e0')),
        ('TEXTCOLOR', (2,1), (2,1), colors.red),
        ('TEXTCOLOR', (2,2), (2,2), colors.orange),
        ('TEXTCOLOR', (2,3), (2,3), colors.HexColor('#b7950b')), # Dark yellow
        ('TEXTCOLOR', (2,4), (2,4), colors.green),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (1,1), (1,-1), 'Helvetica-Bold')
    ]))
    elements.append(map_table)
    elements.append(Spacer(1, 0.2*inch))
    
    # 4. Hybrid Logic Explanation
    elements.append(Paragraph("4. HYBRID LOGIC EXPLANATION", subheading_style))
    logic_text = """
    The system uses a <b>Hybrid Safety-First</b> approach:<br/>
    <b>Step A: Rule Check.</b> Scans for specific keywords. If "cardiac arrest" is found, it immediately sets Level 1 (CRITICAL).<br/>
    <b>Step B: AI Context.</b> If no clear keywords are found, Artificial Intelligence analyzes the context (e.g. distinguishing negation "No chest pain").<br/>
    <b>Step C: Safety Net.</b> The system compares the ESI Rule result vs the AI Analysis. <b>It ALWAYS picks the higher urgency level</b> for safety.<br/>
    """
    elements.append(Paragraph(logic_text, normal_style))
    
    # Build PDF
    doc.build(elements)
    
    print(f"   ✅ PDF report generated: {PDF_OUTPUT}\n")
    return str(PDF_OUTPUT)






def generate_excel_report(processed_calls):
    """Generate Excel report for nurse validation with multiple sheets and styling"""
    print("📊 Step 4: Generating Excel validation report...")
    
    # --- SHEET 1: Validation Data ---
    validation_data = []
    
    # --- SHEET 2: Logic & Reasoning ---
    logic_data = []
    
    # --- SHEET 3: Review Form (Simplified) ---
    review_form_data = []
    
    for call in processed_calls:
        # Combine S+O
        so_summary = ""
        if call.get('subjective') and call.get('objective'):
            so_summary = f"{call['subjective']} {call['objective']}"
        elif call.get('subjective'):
            so_summary = call['subjective']
        elif call.get('objective'):
            so_summary = call['objective']
        else:
            so_summary = "[No clinical summary extracted]"
            
        # Data for Sheet 1
        v_row = {
            'Case ID': call['call_id'],
            'Patient Name': call.get('patient_name', 'N/A'),
            'Expected Urgency': call.get('expected_urgency', 'N/A'),
            'AI Predicted Urgency': call.get('urgency_level', 'UNKNOWN'),
            'Clinical Summary (S+O)': so_summary,
            'NURSE CHECK: Triage Accuracy': '', 
            'NURSE CHECK: Notes Accuracy': '',  
            'Nurse Comments': ''
        }
        validation_data.append(v_row)
        
        # Data for Sheet 2
        l_row = {
            'Case ID': call['call_id'],
            'AI Predicted Urgency': call.get('urgency_level', 'UNKNOWN'),
            'AI Triage Reasoning / Logic': call.get('urgency_reasoning', ''),
            'Assessment & Plan (Context)': f"Assessment: {call.get('assessment','N/A')}\nPlan: {call.get('plan','N/A')}"
        }
        logic_data.append(l_row)
        
        # Data for Sheet 3- The Review Form
        r_row = {
            'Case ID': call['call_id'],
            'Patient': call.get('patient_name', 'N/A'),
            'Notes Quality (Accurate/Missing Info)': '', # Input
            'If Missing, WHAT is missing?': '',          # Input
            'Triage Assessment (Agree/Disagree)': '',    # Input
            'If Disagree, Correct Level?': '',           # Input
            'Final Review Comments': ''                  # Input
        }
        review_form_data.append(r_row)
        
    df_val = pd.DataFrame(validation_data)
    df_logic = pd.DataFrame(logic_data)
    df_review = pd.DataFrame(review_form_data)
    

    # --- SHEET 4: System Logic Explanation (Static Text) ---
    explanation_data = [
        {'Topic': 'SYSTEM OVERVIEW', 'Details': 'The AI uses a Hybrid Safety-First approach combining clinical rules (ESI) with context norms.'},
        {'Topic': '', 'Details': ''},
        {'Topic': 'SCIENTIFIC BASIS', 'Details': 'Based on Emergency Severity Index (ESI) Version 4/5.'},
        {'Topic': 'Reference', 'Details': 'Gilboy N, Tanabe P, Travers D, Rosenau AM. Emergency Severity Index (ESI): A Triage Tool for Emergency Department Care. AHRQ Publication No. 12-0014. Rockville, MD.'},
        {'Topic': '', 'Details': ''},
        {'Topic': '1. ESI STANDARD LEVELS', 'Details': 'Standard Medical Triage Scale:'},
        {'Topic': 'ESI Level 1', 'Details': 'Resuscitation (Immediate life saving intervention required).'},
        {'Topic': 'ESI Level 2', 'Details': 'Emergent (High risk, confused/lethargic, severe pain/distress).'},
        {'Topic': 'ESI Level 3', 'Details': 'Urgent (Stable, needs 2+ resources).'},
        {'Topic': 'ESI Level 4', 'Details': 'Less Urgent (Stable, needs 1 resource).'},
        {'Topic': 'ESI Level 5', 'Details': 'Non-Urgent (No resources needed).'},
        {'Topic': '', 'Details': ''},
        {'Topic': '2. SYSTEM MAPPING', 'Details': 'How ESI maps to our Dashboard Logic:'},
        {'Topic': 'ESI Level 1', 'Details': '-> Maps to CRITICAL (Red)'},
        {'Topic': 'ESI Level 2', 'Details': '-> Maps to HIGH PRIORITY (Orange)'},
        {'Topic': 'ESI Level 3', 'Details': '-> Maps to MEDIUM PRIORITY (Yellow)'},
        {'Topic': 'ESI Level 4 & 5', 'Details': '-> Maps to LOW PRIORITY (Green)'},
        {'Topic': '', 'Details': ''},
        {'Topic': '3. HOW IT WORKS', 'Details': 'Hybrid Process:'},
        {'Topic': 'Step A: Rule Check', 'Details': 'Scans for specific keywords (e.g., "cardiac arrest" -> ESI 1, "chest pain" -> ESI 2).'},
        {'Topic': 'Step B: AI Context', 'Details': 'If no clear keywords, AI analyzes symptoms/risk factors to assign a level.'},
        {'Topic': 'Step C: Safety Net', 'Details': 'Always chooses the HIGHER urgency if there is ambiguity.'},
    ]
    df_explanation = pd.DataFrame(explanation_data)
    
    # Create Excel writer with openpyxl engine
    with pd.ExcelWriter(EXCEL_OUTPUT, engine='openpyxl') as writer:
        
        # Write Sheets
        df_val.to_excel(writer, index=False, sheet_name='Triage & Notes Data')
        df_review.to_excel(writer, index=False, sheet_name='NURSE REVIEW FORM')
        df_logic.to_excel(writer, index=False, sheet_name='AI Logic & Reasoning')
        df_explanation.to_excel(writer, index=False, sheet_name='System Logic Explanation')
        
        # --- STYLING ---
        workbook = writer.book
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        highlight_expected = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        highlight_ai = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        input_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # --- Style Sheet 1 (Data) ---
        ws_val = writer.sheets['Triage & Notes Data']
        widths_val = {'A': 15, 'B': 20, 'C': 18, 'D': 18, 'E': 60, 'F': 25, 'G': 25, 'H': 40}
        
        for col, width in widths_val.items():
            ws_val.column_dimensions[col].width = width
            
        for cell in ws_val[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
        for row in range(2, len(validation_data) + 2):
            ws_val.cell(row=row, column=3).fill = highlight_expected
            ws_val.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            c = ws_val.cell(row=row, column=4)
            c.fill = highlight_ai
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            ws_val.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

        # --- Style Sheet 2 (Review Form) ---
        ws_review = writer.sheets['NURSE REVIEW FORM']
        widths_review = {'A': 15, 'B': 18, 'C': 35, 'D': 45, 'E': 30, 'F': 25, 'G': 45}
        
        for col, width in widths_review.items():
            ws_review.column_dimensions[col].width = width
            
        for cell in ws_review[1]:
            cell.fill = PatternFill(start_color="00796B", end_color="00796B", fill_type="solid")
            cell.font = header_font
            
        for row in range(2, len(review_form_data) + 2):
            ws_review.cell(row=row, column=1).font = Font(bold=True)
            for col_idx in range(3, 8):
                cell = ws_review.cell(row=row, column=col_idx)
                cell.fill = input_fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)

        # --- Style Sheet 3 (Logic) ---
        ws_logic = writer.sheets['AI Logic & Reasoning']
        widths_logic = {'A': 15, 'B': 18, 'C': 70, 'D': 50}
        for col, width in widths_logic.items():
            ws_logic.column_dimensions[col].width = width
            
        for cell in ws_logic[1]:
            cell.fill = header_fill
            cell.font = header_font
            
        for row in range(2, len(logic_data) + 2):
            ws_logic.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
            ws_logic.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
            
        # --- Style Sheet 4 (System Explanation) ---
        ws_expl = writer.sheets['System Logic Explanation']
        ws_expl.column_dimensions['A'].width = 25
        ws_expl.column_dimensions['B'].width = 100
        
        for cell in ws_expl[1]:
            cell.fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
            cell.font = header_font
            
        for row in range(2, len(explanation_data) + 2):
            ws_expl.cell(row=row, column=1).font = Font(bold=True)
            ws_expl.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            
    print(f"   ✅ Excel report generated: {EXCEL_OUTPUT}\n")
    return str(EXCEL_OUTPUT)



def main():
    """Main execution function"""
    print("=" * 70)
    print("SOAP NOTES VALIDATION - Test Data Preparation")
    print("=" * 70)
    print()
    
    # Step 1: Clear all data
    clear_all_data()
    
    # Step 2: Import and process test cases
    processed_calls = import_test_cases()
    
    if not processed_calls:
        print("❌ No calls were processed. Aborting PDF generation.")
        return
    
    # Step 3: Generate PDF report
    pdf_path = generate_pdf_report(processed_calls)
    
    # Step 4: Generate Excel report
    excel_path = generate_excel_report(processed_calls)
    
    print("=" * 70)
    print("✅ PROCESS COMPLETE!")
    print("=" * 70)
    print(f"\n📄 Validation PDF ready: {pdf_path}")
    print(f"📊 Validation Excel ready: {excel_path}")
    print(f"📈 Total cases processed: {len(processed_calls)}/15")
    print("\n📌 Next Steps:")
    print("   1. Open the PDF report or Excel sheet")
    print("   2. Review the dashboard at http://localhost:3000")
    print("   3. Send the files to the nurse for validation")
    print("\n💡 The transparent dashboard shows the S & O summaries")
    print("   for real-time verification during review.")
    print()

if __name__ == "__main__":
    main()

